from openpyxl.styles import Border
import time
import os
import datetime
import pandas as pd
import openpyxl
from sqlalchemy import or_,and_
from config import INDEX,SPIDER_PATH,TEMP_PATH_ONE_MONTH
from utils import excel_operate

from openpyxl.styles import Border, Side,Alignment, PatternFill
from utils.sql_utils import sql_orm
from utils.send_ding_msg import dingmsg
from config import INDEX
from websource.spider.down_foura.foura_spider_universal import serch_performence

class predict_dc():
    def __init__(self):
        self.time_stamp=datetime.datetime.now().strftime('%Y%m%d%H%M')
        self.folder=f'{INDEX}message/nanfen_overtime/xls_dc'
        self.down_path= f"{INDEX}message/battery_life/xls/活动告警.csv"
        self.path_backup = os.path.join(TEMP_PATH_ONE_MONTH,f'{self.time_stamp}.xlsx')
        self.path_picture_xlsx = os.path.join(self.folder,'picture.xlsx')

        self.path_backup_temperate = os.path.join(TEMP_PATH_ONE_MONTH,f'{self.time_stamp}temperate.xlsx')
        self.path_picture_xlsx_temperate = os.path.join(self.folder,'picture_temperate.xlsx')

        self.path_picture=f'{TEMP_PATH_ONE_MONTH}{self.time_stamp}DC.png'
        self.path_picture_temperate=f'{TEMP_PATH_ONE_MONTH}{self.time_stamp}DC_temperate.png'

    def temper_list(self,df):
        with sql_orm(database='nanfen').session_scope() as temp:
            sqlsession,Base=temp
            # 生成文本
            pojo_phone = Base.classes.phone_predict# 钉钉电话
            res = sqlsession.query(pojo_phone).filter(pojo_phone.name == '南宁市监控中心').first()
            jkzx_phone=res.phone
            area_list=[]
            num=[]
            men=[]
            for index,row in df.iterrows():
                if row['区县'] not in area_list:
                    area_list.append(row['区县'])
                    res_list=sqlsession.query(pojo_phone).filter(and_(pojo_phone.area==row['区县'],pojo_phone.level=='代维维护主管')).all()
                    for res in res_list:
                        num.append(res.phone)
                        men.append(res.name)
            if num!=[]:
                num.append(jkzx_phone)
                men.append('南宁市监控中心')
            return num,men,area_list


    def add_DC(self):
        sitelist = pd.read_csv(f'{SPIDER_PATH}station\站址信息.csv', usecols=['运维ID', '所属运营商','区县（行政区划）'],dtype=str).rename(columns={'运维ID': '站址运维ID'})

        alarmlist = pd.read_csv(self.down_path, usecols=['市','站址保障等级', '站址名称', '站址运维ID','告警名称', '告警历时(分钟)','设备告警开始时间','告警详情'],dtype=str)
        alarmlist=alarmlist.loc[alarmlist['市']=='南宁市分公司']
        # 区分核心告警和辅助告警
        alarmlist_list=[alarmlist.loc[alarmlist['告警名称'].isin(['一级低压脱离告警','二级低压脱离告警','交流输入停电告警'])],
                        alarmlist.loc[alarmlist['告警名称'].isin(['整流模块故障告警','整流模块通信状态告警','整流模块风扇告警','整流模块温度过高告警','监控模块故障告警'])],
                        alarmlist.loc[alarmlist['告警名称'].isin(['交流输入缺相告警','交流输入电压过低告警','备电时长过短','电池供电告警','放电过流保护','电池保护告警','电池放电不平衡告警','直流输出电压过低告警','总电压过低'])],
                        alarmlist.loc[alarmlist['告警名称'].isin(['温度过高', '温度超高'])]
                        ]
        i = 0
        for alarmlist in alarmlist_list:
            merge = pd.merge(alarmlist, sitelist, on='站址运维ID', how='left')
            merge=merge.rename(columns={'区县（行政区划）':'区县'})
            column_to_move = merge.pop('区县')
            merge.insert(0, '区县', column_to_move)
            merge.fillna('无', inplace=True)
            if i==3:
                df_highlevel_temper = merge[merge['站址保障等级'].str.contains('L4|L3|L1')]
                num, men, area_list = self.temper_list(df_highlevel_temper)

            merge['告警历时(分钟)']=merge['告警历时(分钟)'].astype(int)
            merge['站址保障等级'] = merge['站址保障等级'].str.split('（').str[0]
            merge['设备告警开始时间'] = merge['设备告警开始时间'].str[5:]
            merge = merge.sort_values(['告警历时(分钟)'], ascending=[False])
            if i==0:
              for index, row in merge.iterrows():
                    merge.loc[index, '直流电压'] = serch_performence().serch_performence_by_id(row['站址运维ID'],'0406111001')

            merge = merge.drop(labels='站址运维ID', axis=1)
            merge = merge.sort_values(['区县', '告警历时(分钟)'], ascending=[True, False])
            rank= merge.groupby('区县')['告警历时(分钟)'].rank(method='first', ascending=False).astype(int)
            merge.insert(0,'排名',rank)
            merge=merge.rename(columns={'告警历时(分钟)':'历时-分钟','设备告警开始时间':'开始时间'})
            if i==0:
                result=merge
                length1=len(merge)
            elif i==1:
                result = pd.concat([result, merge])
                length2 = len(merge)
            elif i==2:
                result=pd.concat([result,merge])
            elif i==3:
                merge['告警名称']=merge['告警详情']
                result_temperate = merge

            i+=1
        result=result.drop('告警详情',axis=1)
        result_temperate=result_temperate.drop('告警详情',axis=1)

        result.to_excel(self.path_backup,index=False)
        result.to_excel(self.path_picture_xlsx,index=False)
        result_temperate.to_excel(self.path_backup_temperate,index=False)
        result_temperate.to_excel(self.path_picture_xlsx_temperate,index=False)
        return length1,length2,num,men,area_list
    def set_format(self,path,title,title1,title2,length1,length2):
        title_row1=length1+3
        title_row2=length1+length2+4

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.insert_rows(1)
        sheet.insert_rows(title_row1)
        sheet.insert_rows(title_row2)

        sheet.cell(row=1, column=1, value=title)
        sheet.cell(row=title_row1, column=1, value=title1)
        sheet.cell(row=title_row2, column=1, value=title2)

        sheet.merge_cells('A1:I1')  # 合并第一行
        sheet.merge_cells(f'A{title_row1}:I{title_row1}')  # 合并第一行
        sheet.merge_cells(f'A{title_row2}:I{title_row2}')  # 合并第一行

        title_cell = sheet['A1']  # 获取合并后的单元格
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        title_cell1 = sheet[f'A{title_row1}']  # 获取合并后的单元格
        title_cell1.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        title_cell2 = sheet[f'A{title_row2}']  # 获取合并后的单元格
        title_cell2.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 23
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 17
        sheet.column_dimensions['I'].width = 15

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in sheet.rows:
            for cell in row:
                cell.border = thin_border
        # 设置所有单元格居中
        for row in sheet.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置值为'疑似退服未发电'的单元格背景
        red_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        title_cell.fill = red_fill
        title_cell1.fill = red_fill
        title_cell2.fill = red_fill

        for row in sheet.rows:
            for cell in row:
                if cell.value == '疑似退服未发电':
                    cell.fill = red_fill
        workbook.save(path)  # 保存文件
    def set_format_temperate(self,path,title):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.insert_rows(1)

        sheet.cell(row=1, column=1, value=title)

        sheet.merge_cells('A1:I1')  # 合并第一行

        title_cell = sheet['A1']  # 获取合并后的单元格
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 23
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 17
        sheet.column_dimensions['I'].width = 15

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in sheet.rows:
            for cell in row:
                cell.border = thin_border
        # 设置所有单元格居中
        for row in sheet.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置值为'疑似退服未发电'的单元格背景
        red_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        title_cell.fill = red_fill

        for row in sheet.rows:
            for cell in row:
                if cell.value == '疑似退服未发电':
                    cell.fill = red_fill
        workbook.save(path)  # 保存文件
    def send_excel(self,length1,length2,num,men,area_list):
        title = '截止' +datetime.datetime.now().strftime('%Y-%m-%d %H:%M')+ '停电历时及当前直流电压情况'
        title1='截止' +datetime.datetime.now().strftime('%Y-%m-%d %H:%M')+ '整流模块故障告警类历时及当前直流电压情况'
        title2='截止' +datetime.datetime.now().strftime('%Y-%m-%d %H:%M')+ '停电类告警类历时及当前直流电压情况'

        area_list='、'.join(area_list)
        add_text = f'{title.replace("停电历时","告警")}，请各单位做好提前预判抢修发电，避免站点退服【直流电压值为“疑似退服未发电”的：该站点已离线】。'
        add_text_temperate='截止' +datetime.datetime.now().strftime('%Y-%m-%d %H:%M')+ '高温故障告警情况'
        self.set_format(self.path_backup,title,title1,title2,length1,length2)
        self.set_format(self.path_picture_xlsx,title,title1,title2,length1,length2)
        self.set_format_temperate(self.path_backup_temperate,add_text_temperate)
        self.set_format_temperate(self.path_picture_xlsx_temperate,add_text_temperate)

        time.sleep(10)
        excel_operate.capture_excel_range(self.path_picture_xlsx, 'Sheet1', self.path_picture)
        time.sleep(10)
        excel_operate.capture_excel_range(self.path_picture_xlsx_temperate, 'Sheet1', self.path_picture_temperate)
        d=dingmsg()
        # 发送其他告警
        d.picture(d.JUJIAO,'http://qyw.gxtower.cn:800/DC_nanfen?timestap='+self.path_picture, add_text)
        d.text_at(d.JUJIAO,f'下载告警详情http://qyw.gxtower.cn:800/DC_file_nanfen?timestap={self.time_stamp}')

        # 发送温度告警
        d.picture(d.JUJIAO,'http://qyw.gxtower.cn:800/DC_nanfen?timestap='+self.path_picture_temperate, add_text_temperate)
        d.text_at(d.JUJIAO,f'以下区域存在高等级温度过高、超高告警：{area_list}\n下载告警详情http://qyw.gxtower.cn:800/DC_file_nanfen?timestap={self.time_stamp}temperate',num,men)

    def run_thread(self):
        length1,length2,num,men,area_list=self.add_DC()
        self.send_excel(length1,length2,num,men,area_list)

# predict_dc().run_thread()