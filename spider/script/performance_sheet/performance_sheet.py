import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from spider.script.down_foura.foura_spider_universal import (
    performence_by_site_list,
    serch_performence,
)

class performance_sheet:
    SIGNAL_LIST = [
        "0406101001", "0406102001", "0406103001",
        "0406111001", "0418101001", "0418102001",
    ]
    SITE_LIST = [
        '45051200000738','45068100000795','45068100000935','45068100000808',
        '45068100000903','45068100000934','45142200000960','45142200000966',
        '45142200000963','45148100000488','45148100000188','45142400000070',
        '45142300000811','45148100000484','45148100000485','45148100000486',
        '45142300000279','45142300000810','45142300000049','45102500001187',
        '45102600000111','45102600000556'
    ]
    SIG2COL = {
        "0406101001": 9,  # I
        "0406102001": 10,
        "0406103001": 11,
        "0406111001": 12,
        "0418101001": 13,
        "0418102001": 14,
    }
    FSU_PATH = r"F:\newtowerV2\websource\spider_download\fsu_chaxun_all\fsu清单.csv"
    OUTPUT_DIR = "spider/down/performance_sheet"
    COOKIE_USER = 1
    TIMEDELTA = 30

    def __init__(self):
        time_str = datetime.now().strftime("%Y%m%d_%H")
        self.out_path = os.path.join(self.OUTPUT_DIR, f"{time_str}.xlsx")

    def _fetch_main(self) -> pd.DataFrame:
        return performence_by_site_list().main(
            site_list=self.SITE_LIST,
            search_id=",".join(self.SIGNAL_LIST),
            cookie_user=self.COOKIE_USER,
            timedelta=self.TIMEDELTA,
        )

    def _fill_missing(self, df_main: pd.DataFrame) -> pd.DataFrame:
        expect = pd.DataFrame(
            [(s, sig) for s in self.SITE_LIST for sig in self.SIGNAL_LIST],
            columns=["站址运维ID", "信号量ID"],
        )
        missing = expect.merge(
            df_main[["站址运维ID", "信号量ID"]], how="left", indicator=True
        )
        missing = missing[missing["_merge"] == "left_only"][["站址运维ID", "信号量ID"]]

        sp = serch_performence()
        fill_rows = []
        for _, row in missing.iterrows():
            station_id = row["站址运维ID"]
            search_id = row["信号量ID"]
            kpi_val = sp.serch_performence_by_id(station_id, search_id)
            ser = pd.Series(index=df_main.columns, dtype="object")
            ser["站址运维ID"] = station_id
            ser["信号量ID"] = search_id
            ser["实测值"] = kpi_val
            fill_rows.append(ser)

        df_fill = pd.DataFrame(fill_rows).reindex(columns=df_main.columns)
        result = pd.concat([df_main, df_fill], ignore_index=True)
        result["时间"] = pd.to_datetime(result["时间"], errors="coerce")
        result["时间"] = result["时间"].fillna(pd.Timestamp.now())
        return result
    def _make_pivot(self, result: pd.DataFrame) -> pd.DataFrame:
        return result.pivot_table(
            index="站址运维ID",
            columns="信号量ID",
            values="实测值",
            aggfunc="first",
        )

    def _write_excel(self, result: pd.DataFrame, pivot: pd.DataFrame):
        fsu_df = pd.read_csv(
            self.FSU_PATH, usecols=["站址运维ID", "FSU在线状态", "站址状态"], dtype=str
        )
        station_time = (
            result.groupby("站址运维ID")["时间"]
            .max()
            .dt.strftime("%Y-%m-%d %H:%M:%S")
        )
        fsu_online = fsu_df.set_index("站址运维ID")["FSU在线状态"].to_dict()
        site_stat = fsu_df.set_index("站址运维ID")["站址状态"].to_dict()

        wb = load_workbook("22站.xlsx")
        ws = wb.active

        for row in range(4, 26):
            station_id = str(ws.cell(row=row, column=17).value).strip()  # Q
            ws.cell(row=row, column=7, value=fsu_online.get(station_id, ""))  # G
            ws.cell(row=row, column=8, value=site_stat.get(station_id, ""))  # H
            for sig, col_num in self.SIG2COL.items():
                ws.cell(
                    row=row,
                    column=col_num,
                    value=pivot.at[station_id, sig]
                    if station_id in pivot.index
                    else "",
                )
            ws.cell(
                row=row,
                column=15,
                value=station_time.get(
                    station_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ),
            )
        wb.save(self.out_path)
        print(f"[PerformanceSheet] 结果已写入 -> {self.out_path}")

    def run(self):
        df_main = self._fetch_main()
        result = self._fill_missing(df_main)
        pivot = self._make_pivot(result)
        self._write_excel(result, pivot)


if __name__ == "__main__":
    performance_sheet().run()
